# -*- coding: utf-8 -*-
import base64
import io

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class PurchaseOrderExcelImport(models.TransientModel):
    _name = 'purchase.order.excel.import'
    _description = 'Purchase Order Excel Import'

    purchase_id = fields.Many2one(
        'purchase.order', string='Purchase Order', required=True,
        default=lambda self: self.env.context.get('active_id'),
    )
    file_data = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')

    
    def action_import(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_('Python package "openpyxl" is required on the Odoo server.'))
        if self.purchase_id.state not in ('draft', 'sent'):
            raise UserError(_('Excel lines can only be imported into a quotation/RFQ in Draft or Sent state.'))

        try:
            content = base64.b64decode(self.file_data)
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise UserError(_('Unable to read the Excel file: %s') % exc)

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise UserError(_('The Excel file is empty.'))

        headers = [str(value).strip().lower() if value is not None else '' for value in rows[0]]
        required = ['product']
        missing = [name for name in required if name not in headers]
        if missing:
            raise ValidationError(_('Missing required Excel column(s): %s') % ', '.join(missing))

        def col(name):
            return headers.index(name) if name in headers else None

        product_col = col('product')
        ref_col = col('internal reference')
        qty_col = col('quantity')
        uom_col = col('uom')
        price_col = col('unit price')
        taxes_col = col('taxes')
        description_col = col('description')

        Product = self.env['product.product']
        Uom = self.env['uom.uom']
        Tax = self.env['account.tax']
        errors = []
        prepared = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(value not in (None, '') for value in row):
                continue

            product_name = row[product_col] if product_col is not None and product_col < len(row) else None
            internal_ref = row[ref_col] if ref_col is not None and ref_col < len(row) else None
            quantity = row[qty_col] if qty_col is not None and qty_col < len(row) else 1
            uom_name = row[uom_col] if uom_col is not None and uom_col < len(row) else None
            unit_price = row[price_col] if price_col is not None and price_col < len(row) else 0
            tax_value = row[taxes_col] if taxes_col is not None and taxes_col < len(row) else None
            description = row[description_col] if description_col is not None and description_col < len(row) else None

            if not product_name and not internal_ref:
                errors.append(_('Row %s: Product or Internal Reference is required.') % row_number)
                continue

            product = False
            if internal_ref:
                product = Product.search([('default_code', '=', str(internal_ref).strip())], limit=1)
            if not product and product_name:
                product = Product.search([('name', '=', str(product_name).strip())], limit=1)
            if not product and product_name:
                product = Product.search([('name', 'ilike', str(product_name).strip())], limit=1)
            if not product:
                errors.append(_('Row %s: Product "%s" was not found.') % (row_number, internal_ref or product_name))
                continue

            try:
                quantity = float(quantity or 1)
                if quantity <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                errors.append(_('Row %s: Quantity must be a positive number.') % row_number)
                continue

            try:
                unit_price = float(unit_price or 0)
            except (TypeError, ValueError):
                errors.append(_('Row %s: Unit Price must be a number.') % row_number)
                continue

            uom = product.uom_po_id
            if uom_name:
                uom = Uom.search([('name', '=', str(uom_name).strip())], limit=1)
                if not uom:
                    uom = Uom.search([('name', 'ilike', str(uom_name).strip())], limit=1)
                if not uom:
                    errors.append(_('Row %s: UoM "%s" was not found.') % (row_number, uom_name))
                    continue

            taxes = False
            if tax_value:
                tax_names = [x.strip() for x in str(tax_value).split(',') if x.strip()]
                found_taxes = self.env['account.tax']
                for tax_name in tax_names:
                    tax = Tax.search([
                        ('name', '=', tax_name),
                        ('type_tax_use', '=', 'purchase'),
                        ('company_id', '=', self.purchase_id.company_id.id),
                    ], limit=1)
                    if not tax:
                        errors.append(_('Row %s: Purchase tax "%s" was not found.') % (row_number, tax_name))
                    else:
                        found_taxes |= tax
                if errors and any(('Row %s:' % row_number) in e for e in errors):
                    continue
                taxes = found_taxes

            prepared.append({
                'product': product,
                'quantity': quantity,
                'uom': uom,
                'unit_price': unit_price,
                'taxes': taxes,
                'description': description,
            })

        if errors:
            raise ValidationError(_('Excel import failed:\n\n%s') % '\n'.join(errors))
        if not prepared:
            raise ValidationError(_('No valid purchase order lines were found in the Excel file.'))

        for vals in prepared:
            line_vals = {
                'order_id': self.purchase_id.id,
                'product_id': vals['product'].id,
                'name': vals['description'] or vals['product'].get_product_multiline_description_purchase() or vals['product'].display_name,
                'product_qty': vals['quantity'],
                'product_uom': vals['uom'].id,
                'price_unit': vals['unit_price'],
                'date_planned': fields.Datetime.now(),
            }
            if vals['taxes']:
                line_vals['taxes_id'] = [(6, 0, vals['taxes'].ids)]
            self.env['purchase.order.line'].create(line_vals)

        workbook.close()
        return {'type': 'ir.actions.act_window_close'}

    
    def action_download_template(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_('Python package "openpyxl" is required on the Odoo server.'))

        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Purchase Order'
        headers = ['Product', 'Internal Reference', 'Quantity', 'UoM', 'Unit Price', 'Taxes', 'Description']
        sheet.append(headers)
        sheet.append(['Test Product', 'TEST001', 1, 'Units', 500, '18%', 'Purchase line description'])
        sheet.append(['Another Product', 'PROD002', 5, 'Units', 1000, '18%', ''])
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        widths = [30, 20, 12, 15, 15, 15, 40]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
        workbook.save(output)
        workbook.close()
        output.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'purchase_order_import_template.xlsx',
            'datas': base64.b64encode(output.read()),
            'datas_fname': 'purchase_order_import_template.xlsx',
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }
