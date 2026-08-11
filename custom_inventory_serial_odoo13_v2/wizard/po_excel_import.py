# -*- coding: utf-8 -*-
from odoo import models, fields, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
from openpyxl import load_workbook

class PurchaseOrderExcelImport(models.TransientModel):
    _name = 'purchase.order.excel.import'
    _description = 'Purchase Order Excel Import'

    file = fields.Binary(required=True)
    filename = fields.Char()

    def action_import(self):
        if not self.file:
            raise UserError(_('Please upload an Excel file.'))
        wb = load_workbook(filename=BytesIO(base64.b64decode(self.file)), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise UserError(_('The Excel file has no item rows.'))

        vendor_name = rows[0][0]
        if not vendor_name:
            raise UserError(_('Vendor is required in the first data row.'))
        partner = self.env['res.partner'].search([('name', '=', vendor_name)], limit=1)
        if not partner:
            raise UserError(_('Vendor not found: %s') % vendor_name)

        order = self.env['purchase.order'].create({'partner_id': partner.id})
        for row_no, row in enumerate(rows, start=2):
            product_code = row[2]
            qty = row[3] or 0
            price = row[5] or 0
            if not product_code:
                continue
            product = self.env['product.product'].search([('default_code', '=', product_code)], limit=1)
            if not product:
                raise UserError(_('Row %s: Product code not found: %s') % (row_no, product_code))
            self.env['purchase.order.line'].create({
                'order_id': order.id,
                'product_id': product.id,
                'name': product.display_name,
                'product_qty': qty,
                'price_unit': price,
                'product_uom': product.uom_po_id.id or product.uom_id.id,
                'date_planned': fields.Datetime.now(),
            })
        return {'type': 'ir.actions.act_window', 'res_model': 'purchase.order',
                'res_id': order.id, 'view_mode': 'form', 'target': 'current'}
